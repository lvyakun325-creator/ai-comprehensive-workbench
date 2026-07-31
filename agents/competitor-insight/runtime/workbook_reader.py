"""Read local competitor account Excel exports into a stable evidence input."""

from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from metrics import parse_metric, parse_publish_time


FIELD_ALIASES = {
    "title": ("标题", "文案", "作品标题", "desc"),
    "likes": ("点赞", "点赞数", "likes"),
    "comments": ("评论", "评论数", "comments"),
    "collects": ("收藏", "收藏数", "collects"),
    "shares": ("分享", "分享数", "shares"),
    "publishedAt": ("发布时间", "发布时间戳", "publish_time", "create_time"),
    "url": ("视频链接", "链接", "url"),
}

_ACCOUNT_KEY_ALIASES = {
    "昵称": "nickname",
    "账号昵称": "nickname",
    "账号名称": "nickname",
    "粉丝": "followers",
    "粉丝数": "followers",
    "粉丝数量": "followers",
    "签名": "signature",
    "个人签名": "signature",
    "简介": "signature",
    "sec_uid": "accountId",
    "secuid": "accountId",
    "抖音号": "accountId",
    "unique_id": "accountId",
}

_ACCOUNT_TEXT_LIMITS = {
    "nickname": 200,
    "accountId": 256,
    "signature": 1000,
}


def _normalized_text(value: object) -> str:
    return str(value).strip().casefold() if value is not None else ""


def _field_map(headers: list[object]) -> dict[str, str]:
    normalized_headers = {_normalized_text(header): str(header).strip() for header in headers if header is not None}
    mapping: dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            found = normalized_headers.get(_normalized_text(alias))
            if found:
                mapping[field] = found
                break
    return mapping


def _find_works_sheet(workbook: Any) -> tuple[Any, list[object], dict[str, str]]:
    candidates: list[tuple[int, int, Any, list[object], dict[str, str]]] = []
    for position, sheet in enumerate(workbook.worksheets):
        headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()))
        mapping = _field_map(headers)
        if "title" in mapping:
            candidates.append((len(mapping), -position, sheet, headers, mapping))
    if not candidates:
        raise ValueError("missing_title_field")
    _score, _position, sheet, headers, mapping = max(candidates, key=lambda item: (item[0], item[1]))
    return sheet, headers, mapping


def _read_account(workbook: Any, works_sheet: Any) -> dict[str, object]:
    preferred = next((sheet for sheet in workbook.worksheets if sheet.title == "账号概览"), None)
    sheets = [preferred] if preferred is not None else []
    sheets.extend(sheet for sheet in workbook.worksheets if sheet is not works_sheet and sheet is not preferred)
    if not sheets:
        raise ValueError("missing_account_sheet")

    account: dict[str, object] = {}
    for sheet in sheets:
        for row in sheet.iter_rows(values_only=True):
            key = row[0] if row else None
            value = row[1] if len(row) > 1 else None
            normalized_key = _normalized_text(key)
            if not normalized_key or value is None:
                continue
            field = _ACCOUNT_KEY_ALIASES.get(normalized_key)
            if field and field not in account:
                if field == "followers":
                    account[field] = parse_metric(value)[0]
                else:
                    text = str(value).strip()
                    if not text:
                        continue
                    if len(text) > _ACCOUNT_TEXT_LIMITS[field]:
                        raise ValueError("invalid_account_identity")
                    account[field] = text
    if not account.get("nickname") and not account.get("accountId"):
        raise ValueError("missing_account_identity")
    return account


def read_account_workbook(path: Path | BinaryIO) -> dict[str, object]:
    """Read a single workbook and return the stable account/works input shape."""
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        works_sheet, headers, field_map = _find_works_sheet(workbook)
        header_positions = {_normalized_text(header): index for index, header in enumerate(headers) if header is not None}
        title_index = header_positions[_normalized_text(field_map["title"])]
        missing_fields = [field for field in FIELD_ALIASES if field not in field_map]
        warnings: list[str] = []
        works: list[dict[str, object]] = []

        for row_number, row in enumerate(works_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            raw_title = row[title_index] if title_index < len(row) else None
            if raw_title is None or not str(raw_title).strip():
                continue
            work: dict[str, object] = {"sourceRow": row_number, "title": str(raw_title).strip()}
            for field in ("likes", "comments", "collects", "shares"):
                column = field_map.get(field)
                raw_value = row[header_positions[_normalized_text(column)]] if column else None
                if column and (raw_value is None or (isinstance(raw_value, str) and not raw_value.strip())):
                    warnings.append(f"missing_metric:{field}:row={row_number}")
                parsed_value, field_warnings = parse_metric(raw_value)
                work[field] = parsed_value
                warnings.extend(field_warnings)
            published_column = field_map.get("publishedAt")
            raw_time = row[header_positions[_normalized_text(published_column)]] if published_column else None
            parsed_time, time_warnings = parse_publish_time(raw_time)
            work["publishedAt"] = parsed_time.isoformat() if parsed_time is not None else None
            warnings.extend(time_warnings)
            url_column = field_map.get("url")
            raw_url = row[header_positions[_normalized_text(url_column)]] if url_column else None
            work["url"] = str(raw_url).strip() if raw_url is not None else ""
            work["totalInteractions"] = sum(int(work[field]) for field in ("likes", "comments", "collects", "shares"))
            works.append(work)

        if not works:
            raise ValueError("no_work_rows")
        return {
            "account": _read_account(workbook, works_sheet),
            "works": works,
            "fieldMap": field_map,
            "missingFields": missing_fields,
            "warnings": warnings,
        }
    finally:
        workbook.close()
