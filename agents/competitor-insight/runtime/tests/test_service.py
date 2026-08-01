from concurrent.futures import ThreadPoolExecutor
from datetime import datetime as real_datetime
import io
import json
import os
from pathlib import Path
import re
import sys
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch
import zipfile

from openpyxl import Workbook


RUNTIME_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RUNTIME_DIR))

import service


def workbook_bytes(nickname: str = "测试账号") -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "账号概览"
    overview.append(["昵称", nickname])
    overview.append(["粉丝数", 100])
    overview.append(["签名", "分享日常生活与健康管理常识"])
    works = workbook.create_sheet("作品数据")
    works.append(["标题", "点赞", "评论", "收藏", "分享", "发布时间", "链接"])
    works.append(["第一条作品", 20, 2, 3, 1, "2026-07-01 10:00:00", "https://example.com/1"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def malformed_account_workbook_bytes() -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "账号概览"
    overview.append(["昵称"])
    works = workbook.create_sheet("作品数据")
    works.append(["标题", "点赞"])
    works.append(["仍有作品", 1])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def xlsx_with_compression_bomb() -> bytes:
    source = io.BytesIO(workbook_bytes())
    output = io.BytesIO()
    with zipfile.ZipFile(source) as original, zipfile.ZipFile(
        output,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as modified:
        for member in original.infolist():
            modified.writestr(member, original.read(member.filename))
        modified.writestr("xl/compression-bomb.bin", b"A" * (2 * 1024 * 1024))
    return output.getvalue()


def xlsx_with_member(member_name: str) -> bytes:
    source = io.BytesIO(workbook_bytes())
    output = io.BytesIO()
    with zipfile.ZipFile(source) as original, zipfile.ZipFile(
        output,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as modified:
        for member in original.infolist():
            modified.writestr(member, original.read(member.filename))
        modified.writestr(member_name, b"unexpected")
    return output.getvalue()


def valid_batches(evidence_id: str = "DY-E0001") -> list[dict[str, object]]:
    empty = {
        "claims": [],
        "topicDirections": [],
        "filmingTemplates": [],
        "conversionItems": [],
        "executionDays": [],
    }
    topics = [
        {
            "title": f"方向{label}",
            "angle": "从标题结构切入",
            "evidenceIds": [evidence_id],
            "complianceNotes": ["不承诺疗效"],
        }
        for label in ("一", "二", "三", "四", "五")
    ]
    templates = [
        {
            "name": f"模板{label}",
            "hook": "用生活场景自然开场",
            "structure": ["提出日常问题", "给出管理提醒"],
            "evidenceIds": [evidence_id],
            "complianceNotes": ["不替代医生建议"],
        }
        for label in ("一", "二", "三")
    ]
    days = [
        {
            "day": day,
            "action": "整理素材并完成发布复盘",
            "evidenceIds": [evidence_id],
            "complianceNotes": ["避免夸大宣传"],
        }
        for day in range(1, 8)
    ]
    return [
        {"batchId": "strategy", **empty},
        {"batchId": "performance", **empty},
        {
            "batchId": "execution",
            "claims": [],
            "topicDirections": topics,
            "filmingTemplates": templates,
            "conversionItems": [
                {
                    "action": "提供健康档案和用药提醒服务",
                    "evidenceIds": [evidence_id],
                    "complianceNotes": ["不替代诊疗"],
                }
            ],
            "executionDays": days,
        },
    ]


def valid_content_batch(evidence_id: str) -> dict[str, object]:
    evidence_fields = {
        "evidenceIds": [evidence_id],
        "complianceNotes": ["不承诺疗效"],
    }
    return {
        "batchId": "content",
        "claims": [
            {
                "section": section,
                "statement": statement,
                "strength": "hypothesis",
                "rationale": rationale,
                "verificationPlan": plan,
                **evidence_fields,
            }
            for section, statement, rationale, plan in (
                ("content-overview", "标题呈现日常管理主题", "来自标题字段的有限观察", "补充同主题内容样本后核验"),
                ("content-structure", "画面结构未提供，作为待验证假设", "证据包未提供画面字段", "补充画面记录后核验"),
                ("interaction", "互动数据可作为后续观察信号", "仅来自单条内容数据", "结合更多内容核验"),
                ("conversion", "转化链路未提供，作为待验证假设", "输入未提供商品或私域字段", "补充承接记录后核验"),
            )
        ],
        "topicDirections": [
            {
                "title": f"复用角度{label}",
                "angle": "从日常管理场景切入",
                "strength": "hypothesis",
                "verificationPlan": "补充同类内容样本后核验",
                **evidence_fields,
            }
            for label in ("一", "二", "三")
        ],
        "filmingTemplates": [{
            "name": "单内容拍法",
            "hook": "用主题自然开场",
            "structure": ["主题", "日常提醒"],
            "strength": "hypothesis",
            "verificationPlan": "补充画面与口播记录后核验",
            **evidence_fields,
        }],
        "conversionItems": [{
            "action": "提供合规资料记录入口",
            "strength": "hypothesis",
            "verificationPlan": "补充承接记录后核验",
            **evidence_fields,
        }],
        "executionDays": [],
    }


class ServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = TemporaryDirectory()
        self.project_root = Path(self.temporary_directory.name)
        self.project_patch = patch.object(service, "PROJECT_ROOT", self.project_root)
        self.project_patch.start()

    def tearDown(self) -> None:
        self.project_patch.stop()
        self.temporary_directory.cleanup()

    def _assert_invalid_evidence(self, operation: object) -> None:
        try:
            operation()
        except Exception as error:
            self.assertIs(type(error), ValueError)
            self.assertEqual(str(error), "invalid_evidence_bundle")
        else:
            self.fail("invalid persisted evidence was accepted")

    def _douyin_root(self) -> Path:
        path = self.project_root / "outputs" / "competitor-insight" / "douyin"
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _artifact_request(self, *, task_id: str = "competitor-20260801-a1") -> dict[str, object]:
        task_dir = self.project_root / "outputs" / "competitor-insight" / "douyin" / task_id
        task_dir.mkdir(parents=True, exist_ok=True)
        data_path = task_dir / "结构化数据.json"
        data_path.write_text(json.dumps({
            "status": "success",
            "data": {
                "profile": {"nickname": "抓取账号", "sec_uid": "public-id"},
                "videos": [{
                    "desc": "公开作品",
                    "statistics": {"digg_count": 20, "comment_count": 2, "collect_count": 3, "share_count": 1},
                    "create_time": "2026-07-01 10:00:00",
                    "share_url": "https://example.com/1",
                }],
            },
        }, ensure_ascii=False), encoding="utf-8")
        return {
            "taskId": task_id,
            "platformId": "douyin",
            "inputKind": "account",
            "outputDir": str(task_dir.resolve()),
            "dataPath": str(data_path.resolve()),
            "excelPath": None,
        }

    def _artifact_request_with_31_ranked_items(self) -> dict[str, object]:
        request = self._artifact_request(task_id="competitor-20260801-ranked31")
        data_path = Path(str(request["dataPath"]))
        payload = json.loads(data_path.read_text(encoding="utf-8"))
        payload["data"]["videos"] = [
            {
                "desc": f"公开作品 {index}",
                "statistics": {
                    "digg_count": 1_001 - index if index <= 10 else 32 - index,
                    "comment_count": 101 - index if index <= 10 else 0,
                    "collect_count": 101 - index if index <= 10 else 0,
                    "share_count": 101 - index if index <= 10 else 0,
                },
                "create_time": (
                    f"2026-07-{32 - index:02d} 10:00:00"
                    if index <= 23
                    else f"2026-01-{index - 23:02d} 10:00:00"
                ),
                "share_url": f"https://example.com/{index}",
            }
            for index in range(1, 32)
        ]
        data_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        return request

    def _artifact_variant_request(self, platform_id: str, input_kind: str) -> dict[str, object]:
        suffix = f"{platform_id}-{input_kind}"
        task_id = f"competitor-20260801-{suffix}"
        task_dir = self.project_root / "outputs" / "competitor-insight" / platform_id / task_id
        task_dir.mkdir(parents=True, exist_ok=True)
        if platform_id == "douyin" and input_kind == "account":
            payload = {
                "data": {
                    "profile": {"nickname": "抖音账号", "sec_uid": "dy-public"},
                    "videos": [{
                        "desc": "抖音公开作品",
                        "statistics": {"digg_count": 20, "comment_count": 2, "collect_count": 3, "share_count": 1},
                        "create_time": "2026-07-01 10:00:00",
                        "share_url": "https://www.douyin.com/video/1",
                    }],
                },
            }
        elif platform_id == "xiaohongshu" and input_kind == "account":
            payload = {
                "data": {
                    "profile": {"nickname": "小红书账号", "red_id": "xhs-public"},
                    "notes": [{
                        "display_title": "小红书公开笔记",
                        "interact_info": {"likedCount": 20, "commentCount": 2, "collectedCount": 3, "sharedCount": 1},
                        "time": "2026-07-01 10:00:00",
                        "url": "https://www.xiaohongshu.com/explore/1",
                    }],
                },
            }
        elif platform_id == "douyin" and input_kind == "content":
            payload = {
                "data": {
                    "author": {"nickname": "抖音作者", "sec_uid": "dy-author"},
                    "video": {
                        "desc": "抖音单条正文",
                        "statistics": {"digg_count": 20, "comment_count": 2, "collect_count": 3, "share_count": 1},
                        "create_time": "2026-07-01 10:00:00",
                        "share_url": "https://www.douyin.com/video/1",
                        "duration": 15,
                    },
                    "transcription": {"transcript": "抖音抓取转写"},
                },
            }
        else:
            payload = {
                "data": {
                    "author": {"nickname": "小红书作者", "red_id": "xhs-author"},
                    "note": {
                        "display_title": "小红书单篇标题",
                        "content": "小红书单篇正文",
                        "ocr_cleaned_text": "小红书抓取 OCR",
                        "interact_info": {"likedCount": 20, "commentCount": 2, "collectedCount": 3, "sharedCount": 1},
                        "time": "2026-07-01 10:00:00",
                        "url": "https://www.xiaohongshu.com/explore/1",
                        "image_count": 3,
                    },
                },
            }
        data_path = task_dir / "结构化数据.json"
        data_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        return {
            "taskId": task_id,
            "platformId": platform_id,
            "inputKind": input_kind,
            "outputDir": str(task_dir.resolve()),
            "dataPath": str(data_path.resolve()),
            "excelPath": None,
        }

    def test_analyze_artifacts_uses_only_the_exact_task_directory_and_persists_context(self) -> None:
        request = self._artifact_request()

        result = service.analyze_artifacts(request)

        task_dir = Path(str(request["outputDir"]))
        self.assertEqual(result["outputDir"], str(task_dir))
        self.assertEqual(result["platformId"], "douyin")
        self.assertEqual(result["inputKind"], "account")
        self.assertEqual(set(result["batchInputs"]), {"strategy", "performance", "execution"})
        for batch_id, batch_input in result["batchInputs"].items():
            self.assertEqual(batch_input["batchId"], batch_id)
            self.assertEqual(batch_input["allowedEvidenceIds"], ["DY-E0001"])
        self.assertTrue((task_dir / "抓取账号_证据包.json").is_file())
        self.assertTrue((task_dir / f"{result['evidenceId']}.evidence-session.json").is_file())

    def test_analyze_artifacts_rejects_cross_task_paths(self) -> None:
        request = self._artifact_request()
        other = self._artifact_request(task_id="competitor-20260801-b2")
        request["dataPath"] = other["dataPath"]

        with self.assertRaisesRegex(ValueError, "path_not_allowed"):
            service.analyze_artifacts(request)

    def test_analyze_artifacts_rejects_extra_fields_and_wrong_platform_root(self) -> None:
        request = self._artifact_request()
        request["unexpected"] = "must fail"
        with self.assertRaisesRegex(ValueError, "invalid_request_fields"):
            service.analyze_artifacts(request)

        request = self._artifact_request()
        request["outputDir"] = str(
            self.project_root / "outputs" / "competitor-insight" / "xiaohongshu" / request["taskId"]
        )
        with self.assertRaisesRegex(ValueError, "path_not_allowed"):
            service.analyze_artifacts(request)

    def test_task_session_rejects_v1_and_traversal_before_opening_a_file(self) -> None:
        request = self._artifact_request()
        result = service.analyze_artifacts(request)
        session_path = Path(str(request["outputDir"])) / f"{result['evidenceId']}.evidence-session.json"
        session = json.loads(session_path.read_text(encoding="utf-8"))
        session["evidence"]["evidenceVersion"] = "1.0"
        session_path.write_text(json.dumps(session), encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "invalid_evidence_bundle"):
            service.validate_batch(str(result["evidenceId"]), valid_batches()[0], str(request["outputDir"]))
        with patch.object(service.os, "open", wraps=os.open) as opened:
            with self.assertRaisesRegex(ValueError, "invalid_evidence_id"):
                service.validate_batch("../foreign", valid_batches()[0], str(request["outputDir"]))
        self.assertFalse(any("foreign" in str(call.args[0]) for call in opened.call_args_list))

    def test_task_session_rejects_incomplete_and_untrusted_v2_evidence(self) -> None:
        request = self._artifact_request()
        result = service.analyze_artifacts(request)
        session_path = Path(str(request["outputDir"])) / f"{result['evidenceId']}.evidence-session.json"
        original = json.loads(session_path.read_text(encoding="utf-8"))
        mutations = (
            lambda value: value["evidence"].pop("subject"),
            lambda value: value["evidence"]["items"].append(dict(value["evidence"]["items"][0])),
            lambda value: value["trustedBatchContexts"][0].update({"allowedEvidenceIds": ["DY-E0001"] * 31}),
        )
        for mutate in mutations:
            session = json.loads(json.dumps(original))
            mutate(session)
            session_path.write_text(json.dumps(session), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "invalid_evidence_bundle"):
                service.validate_batch(str(result["evidenceId"]), valid_batches()[0], str(request["outputDir"]))

    def test_task_session_rejects_mutated_context_and_nested_keys(self) -> None:
        request = self._artifact_request()
        result = service.analyze_artifacts(request)
        session_path = Path(str(request["outputDir"])) / f"{result['evidenceId']}.evidence-session.json"
        original = json.loads(session_path.read_text(encoding="utf-8"))
        for mutate in (
            lambda value: value["trustedBatchContexts"][0]["dangerous"].__class__,
            lambda value: value["evidence"]["account"].update({"rawComments": "forbidden"}),
        ):
            session = json.loads(json.dumps(original))
            if mutate.__code__.co_consts and "dangerous" in str(mutate.__code__.co_consts):
                session["trustedBatchContexts"][0]["dangerous"] = True
            else:
                mutate(session)
            session_path.write_text(json.dumps(session), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "invalid_evidence_bundle"):
                service.validate_batch(str(result["evidenceId"]), valid_batches()[0], str(request["outputDir"]))

    def test_task_session_rejects_rankings_and_contexts_mutated_together_for_validate_and_assemble(self) -> None:
        """Would fail if a strategy-only item can be reassigned into execution by editing both persisted operands."""
        request = self._artifact_request_with_31_ranked_items()
        result = service.analyze_artifacts(request)
        evidence_id = str(result["evidenceId"])
        output_dir = str(request["outputDir"])
        session_path = Path(output_dir) / f"{evidence_id}.evidence-session.json"
        session = json.loads(session_path.read_text(encoding="utf-8"))

        self.assertIn("DY-E0028", session["trustedBatchContexts"][0]["allowedEvidenceIds"])
        self.assertNotIn("DY-E0028", session["trustedBatchContexts"][2]["allowedEvidenceIds"])
        session["evidence"]["rankings"]["overall"]["rows"] = [28]
        session["trustedBatchContexts"] = [
            {
                "batchId": "strategy",
                "allowedEvidenceIds": ["DY-E0028", "DY-E0024", "DY-E0025", "DY-E0026", "DY-E0027"],
            },
            {
                "batchId": "performance",
                "allowedEvidenceIds": [
                    "DY-E0028", "DY-E0024", "DY-E0025", "DY-E0026", "DY-E0027",
                    "DY-E0001", "DY-E0002", "DY-E0003", "DY-E0004", "DY-E0005",
                ],
            },
            {
                "batchId": "execution",
                "allowedEvidenceIds": [
                    "DY-E0028", "DY-E0001", "DY-E0002", "DY-E0003", "DY-E0004", "DY-E0005",
                ],
            },
        ]
        session_path.write_text(json.dumps(session, ensure_ascii=False), encoding="utf-8")
        batches = valid_batches("DY-E0028")

        with self.assertRaisesRegex(ValueError, r"^invalid_evidence_bundle$"):
            service.validate_batch(evidence_id, batches[2], output_dir)
        with self.assertRaisesRegex(ValueError, r"^invalid_evidence_bundle$"):
            service.assemble(evidence_id, batches, output_dir)

    def test_task_session_rejects_closed_schema_type_and_bound_mutations_at_every_account_layer(self) -> None:
        request = self._artifact_request()
        result = service.analyze_artifacts(request)
        evidence_id = str(result["evidenceId"])
        output_dir = str(request["outputDir"])
        session_path = Path(output_dir) / f"{evidence_id}.evidence-session.json"
        original = json.loads(session_path.read_text(encoding="utf-8"))
        batches = valid_batches()
        mutations = (
            ("session_unknown", lambda value: value.update({"unexpected": True})),
            ("session_dangerous", lambda value: value.update({"__proto__": {}})),
            ("canonical_unknown", lambda value: value["canonicalInput"].update({"unexpected": True})),
            ("canonical_source_dangerous", lambda value: value["canonicalInput"]["source"].update({"prototype": {}})),
            ("canonical_parsed_unknown", lambda value: value["canonicalInput"]["parsed"].update({"raw": {}})),
            ("canonical_work_unknown", lambda value: value["canonicalInput"]["parsed"]["works"][0].update({"constructor": {}})),
            ("evidence_dangerous", lambda value: value["evidence"].update({"constructor": {}})),
            ("subject_unknown", lambda value: value["evidence"]["subject"].update({"rawProfile": "forbidden"})),
            ("subject_dangerous", lambda value: value["evidence"]["subject"].update({"prototype": {}})),
            ("account_unknown", lambda value: value["evidence"]["account"].update({"rawComments": []})),
            ("source_unknown", lambda value: value["evidence"]["source"].update({"cookie": "forbidden"})),
            ("source_wrong_type", lambda value: value["evidence"].update({"source": "not-an-object"})),
            ("completeness_unknown", lambda value: value["evidence"]["completeness"].update({"dangerous": True})),
            ("completeness_wrong_type", lambda value: value["evidence"].update({"completeness": []})),
            ("field_map_unknown", lambda value: value["evidence"]["completeness"]["fieldMap"].update({"raw": "raw"})),
            ("field_map_wrong_type", lambda value: value["evidence"]["completeness"].update({"fieldMap": []})),
            ("missing_fields_oversized", lambda value: value["evidence"]["completeness"].update({"missingFields": ["url"] * 501})),
            ("warnings_oversized_text", lambda value: value["evidence"]["completeness"].update({"warnings": ["x" * 100_000]})),
            ("warnings_oversized_array", lambda value: value["evidence"]["completeness"].update({"warnings": ["warning"] * 5_001})),
            ("availability_unknown", lambda value: value["evidence"]["completeness"]["availability"].update({"prototype": False})),
            ("availability_wrong_type", lambda value: value["evidence"]["completeness"]["availability"].update({"comments": 1})),
            ("metrics_unknown", lambda value: value["evidence"]["metrics"].update({"dangerous": 1})),
            ("metrics_wrong_type", lambda value: value["evidence"]["metrics"].update({"workCount": "not-a-number"})),
            ("rankings_unknown", lambda value: value["evidence"]["rankings"].update({"dangerous": {}})),
            ("ranking_wrong_type", lambda value: value["evidence"]["rankings"].update({"overall": []})),
            ("ranking_unknown", lambda value: value["evidence"]["rankings"]["overall"].update({"__proto__": {}})),
            ("ranking_rows_wrong_type", lambda value: value["evidence"]["rankings"]["overall"].update({"rows": "not-a-list"})),
            ("ranking_rows_oversized", lambda value: value["evidence"]["rankings"]["overall"].update({"rows": list(range(1, 12))})),
            ("items_wrong_type", lambda value: value["evidence"].update({"items": {}})),
            ("items_oversized", lambda value: value["evidence"].update({"items": value["evidence"]["items"] * 501})),
            ("item_unknown", lambda value: value["evidence"]["items"][0].update({"dangerous": True})),
            ("item_title_wrong_type", lambda value: value["evidence"]["items"][0].update({"title": {"object": "not-text"}})),
            ("item_title_oversized", lambda value: value["evidence"]["items"][0].update({"title": "x" * 100_000})),
            ("item_metric_wrong_type", lambda value: value["evidence"]["items"][0].update({"likes": True})),
            ("item_url_oversized", lambda value: value["evidence"]["items"][0].update({"url": "https://example.com/" + "x" * 100_000})),
            ("item_ranks_unknown", lambda value: value["evidence"]["items"][0]["ranks"].update({"prototype": 1})),
            ("item_rank_wrong_type", lambda value: value["evidence"]["items"][0]["ranks"].update({"overall": "first"})),
        )
        for name, mutate in mutations:
            with self.subTest(name=name):
                session = json.loads(json.dumps(original))
                mutate(session)
                session_path.write_text(json.dumps(session, ensure_ascii=False), encoding="utf-8")
                self._assert_invalid_evidence(
                    lambda: service.validate_batch(evidence_id, batches[0], output_dir)
                )
                self._assert_invalid_evidence(
                    lambda: service.assemble(evidence_id, batches, output_dir)
                )

    def test_content_task_session_rejects_open_wrong_type_oversized_and_extra_item_shapes(self) -> None:
        request = self._artifact_variant_request("xiaohongshu", "content")
        result = service.analyze_artifacts(request)
        evidence_id = str(result["evidenceId"])
        output_dir = str(request["outputDir"])
        session_path = Path(output_dir) / f"{evidence_id}.evidence-session.json"
        original = json.loads(session_path.read_text(encoding="utf-8"))
        batch = valid_content_batch("XHS-E0001")
        mutations = (
            ("content_unknown", lambda value: value["evidence"]["content"].update({"dangerous": True})),
            ("canonical_content_unknown", lambda value: value["canonicalInput"]["parsed"]["content"].update({"dangerous": True})),
            ("content_dangerous", lambda value: value["evidence"]["content"].update({"constructor": {}})),
            ("content_wrong_type", lambda value: value["evidence"].update({"content": "not-an-object"})),
            ("content_body_wrong_type", lambda value: value["evidence"]["content"].update({"body": {"not": "text"}})),
            ("content_body_oversized", lambda value: value["evidence"]["content"].update({"body": "x" * 100_000})),
            ("content_author_unknown", lambda value: value["evidence"]["content"]["author"].update({"rawProfile": "forbidden"})),
            ("content_author_wrong_type", lambda value: value["evidence"]["content"].update({"author": []})),
            ("content_numeric_wrong_type", lambda value: value["evidence"]["content"].update({"imageCount": True})),
            ("content_rankings_nonempty", lambda value: value["evidence"].update({"rankings": {"overall": {"status": "available", "rows": [1]}}})),
            ("content_extra_item", lambda value: value["evidence"]["items"].append(dict(value["evidence"]["items"][0]))),
        )
        for name, mutate in mutations:
            with self.subTest(name=name):
                session = json.loads(json.dumps(original))
                mutate(session)
                session_path.write_text(json.dumps(session, ensure_ascii=False), encoding="utf-8")
                self._assert_invalid_evidence(
                    lambda: service.validate_batch(evidence_id, batch, output_dir)
                )
                self._assert_invalid_evidence(
                    lambda: service.assemble(evidence_id, [batch], output_dir)
                )

    def test_all_four_artifact_variants_validate_and_assemble_from_their_task_session(self) -> None:
        for platform_id, input_kind, evidence_id in (
            ("douyin", "account", "DY-E0001"),
            ("xiaohongshu", "account", "XHS-E0001"),
            ("douyin", "content", "DY-E0001"),
            ("xiaohongshu", "content", "XHS-E0001"),
        ):
            with self.subTest(platform_id=platform_id, input_kind=input_kind):
                request = self._artifact_variant_request(platform_id, input_kind)
                result = service.analyze_artifacts(request)
                batches = valid_batches(evidence_id) if input_kind == "account" else [valid_content_batch(evidence_id)]
                validated = service.validate_batch(
                    str(result["evidenceId"]),
                    batches[-1],
                    str(request["outputDir"]),
                )
                artifact = service.assemble(
                    str(result["evidenceId"]),
                    batches,
                    str(request["outputDir"]),
                )
                self.assertEqual(validated["stage"], "section_validated")
                self.assertEqual(artifact["stage"], "report_ready")

    def test_rejects_paths_outside_the_controlled_douyin_directory(self) -> None:
        outside = self.project_root / "private.xlsx"
        outside.write_bytes(workbook_bytes())

        with self.assertRaisesRegex(ValueError, "path_outside_douyin_output"):
            service.analyze_path(str(outside))
        with self.assertRaisesRegex(ValueError, "path_outside_douyin_output"):
            service.analyze_path(str(self._douyin_root() / ".." / ".." / "private.xlsx"))

    def test_rejects_a_symlink_before_resolving_its_external_target(self) -> None:
        outside = self.project_root / "private.xlsx"
        outside.write_bytes(workbook_bytes())
        link = self._douyin_root() / "linked.xlsx"
        link.symlink_to(outside)

        with self.assertRaisesRegex(ValueError, "symlink_not_allowed"):
            service.analyze_path(str(link))

    def test_rejects_a_symlink_in_the_controlled_root_ancestry(self) -> None:
        external_outputs = self.project_root / "external-outputs"
        douyin = external_outputs / "competitor-insight" / "douyin"
        douyin.mkdir(parents=True)
        workbook_path = douyin / "account.xlsx"
        workbook_path.write_bytes(workbook_bytes())
        (self.project_root / "outputs").symlink_to(
            external_outputs,
            target_is_directory=True,
        )

        with self.assertRaisesRegex(ValueError, "symlink_not_allowed"):
            service.analyze_path(str(self.project_root / "outputs" / "competitor-insight" / "douyin" / "account.xlsx"))

    def test_reads_from_opened_snapshot_if_checked_path_is_swapped_for_a_symlink(self) -> None:
        workbook_path = self._douyin_root() / "account.xlsx"
        workbook_path.write_bytes(workbook_bytes("原始账号"))
        external = self.project_root / "outside.xlsx"
        external.write_bytes(workbook_bytes("外部账号"))
        real_reader = service.read_account_workbook
        swapped = False

        def swap_then_read(path: Path) -> dict[str, object]:
            nonlocal swapped
            if not swapped:
                workbook_path.unlink()
                workbook_path.symlink_to(external)
                swapped = True
            return real_reader(path)

        with patch.object(service, "read_account_workbook", side_effect=swap_then_read):
            result = service.analyze_path(str(workbook_path))

        self.assertEqual(result["account"]["nickname"], "原始账号")

    def test_rejects_file_swapped_to_symlink_between_lstat_and_open(self) -> None:
        workbook_path = self._douyin_root() / "account.xlsx"
        workbook_path.write_bytes(workbook_bytes())
        external = self.project_root / "outside.xlsx"
        external.write_bytes(workbook_bytes("外部账号"))
        real_open = os.open
        swapped = False

        def swap_before_open(
            path: str | Path,
            flags: int,
            mode: int = 0o777,
            *,
            dir_fd: int | None = None,
        ) -> int:
            nonlocal swapped
            if path == "account.xlsx" and dir_fd is not None and not swapped:
                workbook_path.unlink()
                workbook_path.symlink_to(external)
                swapped = True
            return real_open(path, flags, mode, dir_fd=dir_fd)

        with patch.object(service.os, "open", side_effect=swap_before_open):
            with self.assertRaisesRegex(ValueError, "symlink_not_allowed"):
                service.analyze_path(str(workbook_path))

    def test_fails_closed_for_ordinary_path_when_nofollow_is_zero(self) -> None:
        workbook_path = self._douyin_root() / "account.xlsx"
        workbook_path.write_bytes(workbook_bytes())

        with patch.object(service.os, "O_NOFOLLOW", 0):
            with self.assertRaisesRegex(ValueError, r"^secure_nofollow_unavailable$"):
                service.analyze_path(str(workbook_path))

    def test_fails_closed_with_stable_error_when_directory_open_is_unavailable(self) -> None:
        workbook_path = self._douyin_root() / "account.xlsx"
        workbook_path.write_bytes(workbook_bytes())

        with patch.object(service.os, "O_DIRECTORY", 0):
            with self.assertRaisesRegex(ValueError, r"^secure_directory_unavailable$"):
                service.analyze_path(str(workbook_path))

    def test_fails_closed_before_path_swap_when_nofollow_is_missing(self) -> None:
        workbook_path = self._douyin_root() / "account.xlsx"
        workbook_path.write_bytes(workbook_bytes())
        external = self.project_root / "outside.xlsx"
        external.write_bytes(workbook_bytes("外部账号"))
        real_open = os.open

        def swap_if_opened(
            path: str | Path,
            flags: int,
            mode: int = 0o777,
            *,
            dir_fd: int | None = None,
        ) -> int:
            if path == "account.xlsx" and dir_fd is not None:
                workbook_path.unlink()
                workbook_path.symlink_to(external)
            return real_open(path, flags, mode, dir_fd=dir_fd)

        nofollow = service.os.O_NOFOLLOW
        delattr(service.os, "O_NOFOLLOW")
        try:
            with patch.object(service.os, "open", side_effect=swap_if_opened):
                with self.assertRaisesRegex(ValueError, r"^secure_nofollow_unavailable$"):
                    service.analyze_path(str(workbook_path))
        finally:
            service.os.O_NOFOLLOW = nofollow

    def test_fails_closed_when_nofollow_value_cannot_be_used(self) -> None:
        workbook_path = self._douyin_root() / "account.xlsx"
        workbook_path.write_bytes(workbook_bytes())

        with patch.object(service.os, "O_NOFOLLOW", 1 << 100):
            with self.assertRaisesRegex(ValueError, r"^secure_nofollow_unavailable$"):
                service.analyze_path(str(workbook_path))

    def test_rejects_invalid_upload_extension_signature_and_excel_size(self) -> None:
        valid = workbook_bytes()

        with self.assertRaisesRegex(ValueError, "invalid_xlsx_signature"):
            service.analyze_upload("fake.xlsx", b"not-a-zip")
        with self.assertRaisesRegex(ValueError, "invalid_extension"):
            service.analyze_upload("fake.xls", valid)
        with patch.object(service, "MAX_EXCEL_BYTES", len(valid) - 1):
            with self.assertRaisesRegex(ValueError, "excel_too_large"):
                service.analyze_upload("large.xlsx", valid)

    def test_rejects_high_expansion_xlsx_before_decompression(self) -> None:
        with self.assertRaisesRegex(ValueError, "xlsx_archive_too_large"):
            service.analyze_upload("bomb.xlsx", xlsx_with_compression_bomb())

    def test_rejects_backslash_parent_drive_and_unc_xlsx_member_paths(self) -> None:
        for member_name in (
            r"..\outside.txt",
            "xl/./hidden.txt",
            r"C:\outside.txt",
            r"\\server\share.txt",
        ):
            with self.subTest(member_name=member_name):
                with self.assertRaisesRegex(ValueError, "invalid_xlsx_signature"):
                    service.analyze_upload(
                        "unsafe-member.xlsx",
                        xlsx_with_member(member_name),
                    )

    def test_upload_persists_only_id_named_evidence_and_removes_temporary_copy(self) -> None:
        result = service.analyze_upload("sample.xlsx", workbook_bytes())

        evidence_id = result["evidenceId"]
        evidence_path = (
            self.project_root
            / "outputs"
            / "competitor-insight"
            / "reports"
            / "evidence"
            / f"{evidence_id}.json"
        )
        self.assertRegex(str(evidence_id), r"^[0-9a-f]{16}$")
        self.assertTrue(evidence_path.is_file())
        self.assertEqual(json.loads(evidence_path.read_text(encoding="utf-8"))["evidenceId"], evidence_id)
        temporary_root = evidence_path.parents[1] / ".tmp"
        self.assertEqual(list(temporary_root.iterdir()), [])
        self.assertEqual(result["stage"], "evidence_ready")
        self.assertEqual(result["account"]["nickname"], "测试账号")

    def test_output_roots_reject_symlinks_without_writing_outside(self) -> None:
        for component in ("reports", "evidence", ".tmp"):
            with self.subTest(component=component), TemporaryDirectory() as outside_dir:
                outside = Path(outside_dir)
                base = self.project_root / "outputs" / "competitor-insight"
                reports = base / "reports"
                if component == "reports":
                    base.mkdir(parents=True, exist_ok=True)
                    reports.symlink_to(outside, target_is_directory=True)
                else:
                    reports.mkdir(parents=True, exist_ok=True)
                    candidate = reports / component
                    if candidate.is_dir() and not candidate.is_symlink():
                        candidate.rmdir()
                    (reports / component).symlink_to(outside, target_is_directory=True)

                try:
                    with self.assertRaisesRegex(ValueError, r"^unsafe_output_path$"):
                        service.analyze_upload("sample.xlsx", workbook_bytes())
                    self.assertEqual(list(outside.iterdir()), [])
                finally:
                    link = reports if component == "reports" else reports / component
                    if link.is_symlink():
                        link.unlink()

    def test_output_component_swap_to_symlink_is_rejected_without_external_write(self) -> None:
        outside = self.project_root / "outside-output"
        outside.mkdir()
        base = self.project_root / "outputs" / "competitor-insight"
        base.mkdir(parents=True)
        reports = base / "reports"
        reports.mkdir()
        real_open = os.open
        swapped = False

        def swap_before_open(
            path: str | Path,
            flags: int,
            mode: int = 0o777,
            *,
            dir_fd: int | None = None,
        ) -> int:
            nonlocal swapped
            if path == "reports" and dir_fd is not None and not swapped:
                reports.rmdir()
                reports.symlink_to(outside, target_is_directory=True)
                swapped = True
            return real_open(path, flags, mode, dir_fd=dir_fd)

        with patch.object(service.os, "open", side_effect=swap_before_open):
            with self.assertRaisesRegex(ValueError, r"^unsafe_output_path$"):
                service.analyze_upload("sample.xlsx", workbook_bytes())
        self.assertEqual(list(outside.iterdir()), [])

    def test_output_writes_fail_closed_without_directory_open_support(self) -> None:
        with patch.object(service.os, "O_DIRECTORY", 0):
            with self.assertRaisesRegex(ValueError, r"^unsafe_output_path$"):
                service.analyze_upload("sample.xlsx", workbook_bytes())

    def test_evidence_ready_returns_bounded_deterministic_batch_inputs(self) -> None:
        first = service.analyze_upload("sample.xlsx", workbook_bytes())
        second = service.analyze_upload("sample.xlsx", workbook_bytes())

        self.assertEqual(first["batchInputs"], second["batchInputs"])
        self.assertEqual(
            set(first["batchInputs"]),
            {"strategy", "performance", "execution"},
        )
        encoded = json.dumps(first["batchInputs"], ensure_ascii=False)
        self.assertLess(len(encoded.encode("utf-8")), 80_000)
        self.assertNotIn("sample.xlsx", encoded)
        self.assertNotIn("contentBase64", encoded)
        self.assertNotIn("reportPath", encoded)
        for batch_id, batch_input in first["batchInputs"].items():
            with self.subTest(batch_id=batch_id):
                self.assertGreater(len(batch_input["evidence"]), 0)
                self.assertEqual(
                    batch_input["evidence"][0]["evidenceId"],
                    "DY-E0001",
                )
                self.assertNotIn("url", batch_input["evidence"][0])
                self.assertNotIn("sourceRow", batch_input["evidence"][0])
        self.assertEqual(first["batchInputs"]["strategy"]["account"], {
            "nickname": "测试账号",
            "followers": 100,
            "signature": "分享日常生活与健康管理常识",
        })
        self.assertNotIn("account", first["batchInputs"]["performance"])
        self.assertNotIn("account", first["batchInputs"]["execution"])

    def test_analyze_path_reads_without_modifying_the_original_workbook(self) -> None:
        workbook_path = self._douyin_root() / "account.xlsx"
        before = workbook_bytes()
        workbook_path.write_bytes(before)

        result = service.analyze_path(str(workbook_path))

        self.assertEqual(workbook_path.read_bytes(), before)
        self.assertEqual(result["stage"], "evidence_ready")

    def test_validate_batch_loads_the_controlled_evidence_file_on_every_call(self) -> None:
        result = service.analyze_upload("sample.xlsx", workbook_bytes())
        evidence_id = str(result["evidenceId"])
        batch = valid_batches()[0]

        validated = service.validate_batch(evidence_id, batch)
        self.assertEqual(validated["batch"]["batchId"], "strategy")

        evidence_path = (
            self.project_root
            / "outputs"
            / "competitor-insight"
            / "reports"
            / "evidence"
            / f"{evidence_id}.json"
        )
        evidence_path.write_text("{}\n", encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "invalid_evidence_bundle"):
            service.validate_batch(evidence_id, batch)

    def test_rejects_uncontrolled_evidence_ids_and_filename_mismatches(self) -> None:
        with self.assertRaisesRegex(ValueError, "invalid_evidence_id"):
            service.validate_batch("../outside", valid_batches()[0])

        evidence_root = (
            self.project_root / "outputs" / "competitor-insight" / "reports" / "evidence"
        )
        evidence_root.mkdir(parents=True)
        evidence_root.joinpath("0123456789abcdef.json").write_text(
            json.dumps({"evidenceId": "fedcba9876543210"}),
            encoding="utf-8",
        )
        with self.assertRaisesRegex(ValueError, "invalid_evidence_bundle"):
            service.validate_batch("0123456789abcdef", valid_batches()[0])

    def test_assembles_three_validated_batches_and_writes_a_safe_timestamped_report(self) -> None:
        result = service.analyze_upload("sample.xlsx", workbook_bytes("../../坏/账号"))
        evidence_id = str(result["evidenceId"])

        artifact = service.assemble(evidence_id, valid_batches())

        report_path = Path(str(artifact["reportPath"]))
        reports_root = (
            self.project_root / "outputs" / "competitor-insight" / "reports"
        ).resolve()
        self.assertEqual(report_path.parent, reports_root)
        self.assertTrue(report_path.is_file())
        self.assertNotIn("..", report_path.name)
        self.assertNotIn("/", report_path.name)
        self.assertTrue(
            re.fullmatch(
                r"坏_账号_抖音账号分析报告_\d{8}_\d{6}\.md",
                report_path.name,
            )
        )
        self.assertEqual(artifact["validationErrors"], [])
        self.assertEqual(artifact["stage"], "report_ready")
        self.assertEqual(report_path.read_text(encoding="utf-8"), artifact["markdown"])

    def test_assemble_requires_exactly_one_of_each_batch(self) -> None:
        result = service.analyze_upload("sample.xlsx", workbook_bytes())
        evidence_id = str(result["evidenceId"])
        batches = valid_batches()

        with self.assertRaisesRegex(ValueError, "missing_batch_id:performance"):
            service.assemble(evidence_id, [batches[0], batches[2]])
        with self.assertRaisesRegex(ValueError, "duplicate_batch_id:strategy"):
            service.assemble(evidence_id, [batches[0], batches[0], batches[2]])

    def test_same_second_parallel_reports_use_exclusive_distinct_filenames(self) -> None:
        result = service.analyze_upload("sample.xlsx", workbook_bytes())
        evidence_id = str(result["evidenceId"])
        frozen_datetime = unittest.mock.Mock()
        frozen_datetime.now.return_value = real_datetime(2026, 7, 31, 12, 0, 0)

        with patch.object(service, "datetime", frozen_datetime):
            with ThreadPoolExecutor(max_workers=2) as executor:
                artifacts = list(
                    executor.map(
                        lambda _index: service.assemble(evidence_id, valid_batches()),
                        range(2),
                    )
                )

        paths = [Path(str(artifact["reportPath"])) for artifact in artifacts]
        self.assertEqual(len(set(paths)), 2)
        self.assertTrue(any(path.name.endswith("_20260731_120000.md") for path in paths))
        self.assertTrue(any(path.name.endswith("_20260731_120000_01.md") for path in paths))
        for path, artifact in zip(paths, artifacts):
            self.assertEqual(path.read_text(encoding="utf-8"), artifact["markdown"])

    def test_unknown_workbook_value_error_is_normalized(self) -> None:
        with patch.object(
            service,
            "read_account_workbook",
            side_effect=ValueError("unexpected-parser-detail"),
        ):
            with self.assertRaisesRegex(ValueError, r"^invalid_workbook$"):
                service.analyze_upload("malformed.xlsx", workbook_bytes())


if __name__ == "__main__":
    unittest.main()
