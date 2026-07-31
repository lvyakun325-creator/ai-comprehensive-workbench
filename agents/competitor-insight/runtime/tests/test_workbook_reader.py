from pathlib import Path
import sys
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook


RUNTIME_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RUNTIME_DIR))

from workbook_reader import read_account_workbook


class ReadAccountWorkbookTests(unittest.TestCase):
    def _write_workbook(
        self,
        directory: str,
        overview_rows: list[tuple[object, object]],
        headers: list[str],
        work_rows: list[list[object]],
        overview_name: str = "账号概览",
        works_name: str = "全部作品",
    ) -> Path:
        workbook = Workbook()
        overview = workbook.active
        overview.title = overview_name
        for row in overview_rows:
            overview.append(row)
        works = workbook.create_sheet(works_name)
        works.append(headers)
        for row in work_rows:
            works.append(row)
        path = Path(directory) / "account.xlsx"
        workbook.save(path)
        return path

    def test_reads_standard_sheets_and_chinese_fields(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [("昵称", "测试账号"), ("粉丝数", "1.2w")],
                ["标题", "点赞", "评论", "收藏", "分享", "发布时间", "视频链接"],
                [["第一条作品", "8,000", 12, 3, 4, "2026-07-31 12:30", "https://example.com/1"]],
            )

            parsed = read_account_workbook(path)

        self.assertEqual(parsed["account"]["nickname"], "测试账号")
        self.assertEqual(parsed["works"][0]["title"], "第一条作品")
        self.assertEqual(parsed["fieldMap"]["likes"], "点赞")
        self.assertEqual(parsed["works"][0]["likes"], 8000)

    def test_reads_real_douyin_export_account_shape_with_bounded_context(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [
                    ("", ""),
                    ("", ""),
                    ("昵称", "真实导出形状"),
                    ("sec_uid", "MS4wLjABAAAA-test-account"),
                    ("粉丝数", "1.2w"),
                    ("签名", "记录日常生活与健康管理常识"),
                ],
                ["标题", "点赞", "评论", "收藏", "分享", "发布时间", "视频链接"],
                [["第一条作品", 20, 2, 3, 1, "2026-07-01 10:00", "https://example.com/1"]],
            )

            parsed = read_account_workbook(path)

        self.assertEqual(parsed["account"], {
            "nickname": "真实导出形状",
            "accountId": "MS4wLjABAAAA-test-account",
            "followers": 12000,
            "signature": "记录日常生活与健康管理常识",
        })

    def test_rejects_works_only_and_single_work_templates(self) -> None:
        with TemporaryDirectory() as directory:
            for sheet_name in ("随机数据", "单作品导出"):
                with self.subTest(sheet_name=sheet_name):
                    workbook = Workbook()
                    sheet = workbook.active
                    sheet.title = sheet_name
                    sheet.append(["标题", "点赞", "评论", "发布时间"])
                    sheet.append(["只有作品", 10, 2, "2026-07-31 12:30"])
                    path = Path(directory) / f"{sheet_name}.xlsx"
                    workbook.save(path)
                    workbook.close()

                    with self.assertRaisesRegex(ValueError, r"^missing_account_sheet$"):
                        read_account_workbook(path)

    def test_rejects_account_sheet_without_douyin_identity(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [
                    ("昵称", "小红书账号"),
                    ("小红书号", "red-123"),
                    ("签名", "错误平台"),
                ],
                ["标题", "点赞"],
                [["错误平台作品", 10]],
                overview_name="小红书账号信息",
            )

            with self.assertRaisesRegex(ValueError, r"^wrong_platform_account_sheet$"):
                read_account_workbook(path)

    def test_nonstandard_account_sheet_requires_a_douyin_specific_id(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [("昵称", "通用平台账号"), ("粉丝", 99)],
                ["标题", "点赞"],
                [["通用作品", 10]],
                overview_name="资料页",
            )

            with self.assertRaisesRegex(ValueError, r"^missing_account_identity$"):
                read_account_workbook(path)

    def test_rejects_unbounded_account_identity_and_signature(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [("昵称", "账" * 201), ("签名", "签" * 1001)],
                ["标题", "点赞"],
                [["正常作品", 10]],
            )

            with self.assertRaisesRegex(ValueError, r"^invalid_account_identity$"):
                read_account_workbook(path)

    def test_recognizes_english_aliases(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [("账号昵称", "英文账号")],
                ["文案", "likes", "comments", "collects", "shares", "create_time", "url"],
                [["First post", "1.5W", 2, 3, 4, "2026-07-31 12:30", "https://example.com/1"]],
            )

            parsed = read_account_workbook(path)

        self.assertEqual(parsed["account"]["nickname"], "英文账号")
        self.assertEqual(parsed["fieldMap"]["title"], "文案")
        self.assertEqual(parsed["fieldMap"]["publishedAt"], "create_time")
        self.assertEqual(parsed["works"][0]["likes"], 15000)

    def test_reports_missing_collects_and_shares(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [("昵称", "测试账号")],
                ["标题", "点赞", "评论", "发布时间"],
                [["第一条作品", 10, 2, "2026-07-31 12:30"]],
            )

            parsed = read_account_workbook(path)

        self.assertIn("collects", parsed["missingFields"])
        self.assertIn("shares", parsed["missingFields"])
        self.assertEqual(parsed["works"][0]["collects"], 0)
        self.assertEqual(parsed["works"][0]["shares"], 0)

    def test_distinguishes_real_zero_from_missing_metric_cells(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [("昵称", "测试账号")],
                ["标题", "点赞", "评论", "收藏", "分享", "发布时间"],
                [
                    ["真实零互动", 0, 0, 0, 0, "2026-07-31 12:30"],
                    ["缺失互动", None, "", 1, None, "2026-07-31 12:30"],
                ],
            )

            parsed = read_account_workbook(path)

        self.assertEqual(parsed["works"][0]["likes"], 0)
        self.assertEqual(parsed["works"][1]["likes"], 0)
        self.assertIn("missing_metric:likes:row=3", parsed["warnings"])
        self.assertIn("missing_metric:comments:row=3", parsed["warnings"])
        self.assertIn("missing_metric:shares:row=3", parsed["warnings"])
        self.assertNotIn("missing_metric:likes:row=2", parsed["warnings"])

    def test_rejects_missing_title_field(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [("昵称", "测试账号")],
                ["点赞", "评论", "发布时间"],
                [[10, 2, "2026-07-31 12:30"]],
            )

            with self.assertRaisesRegex(ValueError, "missing_title_field"):
                read_account_workbook(path)

    def test_rejects_no_work_rows(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [("昵称", "测试账号")],
                ["标题", "点赞", "评论", "发布时间"],
                [],
            )

            with self.assertRaisesRegex(ValueError, "no_work_rows"):
                read_account_workbook(path)

    def test_discovers_differently_named_sheets(self) -> None:
        with TemporaryDirectory() as directory:
            path = self._write_workbook(
                directory,
                [("昵称", "异名账号"), ("sec_uid", "MS4w-test"), ("粉丝", 99)],
                ["作品标题", "点赞数", "评论数", "收藏数", "分享数", "发布时间戳", "链接"],
                [["异名作品", 10, 2, 3, 4, 1_722_429_000, "https://example.com/1"]],
                overview_name="资料页",
                works_name="内容导出",
            )

            parsed = read_account_workbook(path)

        self.assertEqual(parsed["account"]["nickname"], "异名账号")
        self.assertEqual(parsed["account"]["accountId"], "MS4w-test")
        self.assertEqual(parsed["works"][0]["title"], "异名作品")
        self.assertEqual(parsed["works"][0]["publishedAt"], "2024-07-31T12:30:00")
